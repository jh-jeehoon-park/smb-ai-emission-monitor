#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""산출물 xlsx 3종을 docs/specs/*.md 에서 생성한다.

규칙은 `.claude/rules/deliverable-xlsx.rule.md`. 요점만 옮기면,

* 원천은 마크다운이고 xlsx는 생성물이다(X6). 엑셀에서 직접 고치면 다음 실행 때 사라진다.
* **기존 파일의 셀만 바꾼다.** 발주처 양식의 서식·인쇄 설정·시트 구성을 통째로 새로
  만들면 잃는다. 서식이 openpyxl 왕복을 견디는지는 실측했다(규칙 §7.2).
* **xlsx가 자립해야 한다**(X8). `"상세는 …md 참조"` 같은 칸을 만들지 않고 근거를 글로 넣는다.
* 영역이 아니라 **영역 안의 항목**이 한 행이다(X4).

실행: `python scripts/build_xlsx.py` (프로젝트 루트에서)
"""
from __future__ import annotations

import re
import shutil
import sys
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SPECS = ROOT / "docs/specs"
SCREENS = SPECS / "screens"

XLSX = {
    "screen": SPECS / "오염물질 배출관리_화면설계서_v0.1.xlsx",
    "req": SPECS / "오염물질 배출관리_요구사항정의서_v0.1.xlsx",
    "data": SPECS / "오염물질 배출관리_데이터 정의서_v0.1.xlsx",
}

BUILD_DATE = date.today().isoformat()


# ──────────────────────────── 마크다운 읽기 ────────────────────────────

def read(path: Path) -> str:
    return path.read_text(encoding="utf-8").replace("\r\n", "\n")


ESCAPED_PIPE = "\\" + "|"
PLACEHOLDER = "\x00"


def split_row(line: str) -> list[str]:
    """표 한 줄을 칸 배열로. 타입 유니온의 이스케이프된 파이프는 칸 구분이 아니다."""
    body = line.strip().replace(ESCAPED_PIPE, PLACEHOLDER)
    cells = body.split("|")[1:-1]
    return [c.replace(PLACEHOLDER, "|").strip() for c in cells]


SEPARATOR = re.compile(r"^\|[-: |]+\|$")


def tables(text: str) -> list[tuple[list[str], list[list[str]]]]:
    """문서 안의 모든 표를 (헤더, 행들)로. 헤더 판별은 다음 줄이 구분선인지로 한다."""
    out = []
    lines = text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        nxt = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if line.startswith("|") and SEPARATOR.match(nxt):
            header = split_row(line)
            rows = []
            j = i + 2
            while j < len(lines) and lines[j].strip().startswith("|"):
                rows.append(split_row(lines[j]))
                j += 1
            out.append((header, rows))
            i = j
        else:
            i += 1
    return out


def table_with(text: str, *required: str) -> tuple[list[str], list[list[str]]] | None:
    """지정한 열을 모두 가진 첫 표. **열 이름으로 찾는다** — 표 순서를 가정하지 않는다."""
    for header, rows in tables(text):
        if all(any(c == want for c in header) for want in required):
            return header, rows
    return None


def section(text: str, title_pattern: str) -> str:
    """제목에 해당하는 절의 본문. 다음 같은 수준 이상의 제목 앞까지."""
    m = re.search(r"^(#{2,4}) (" + title_pattern + r").*$", text, re.M)
    if not m:
        return ""
    depth = len(m.group(1))
    start = m.end()
    nxt = re.search(r"^#{1," + str(depth) + r"} ", text[start:], re.M)
    return text[start : start + nxt.start()] if nxt else text[start:]


# ──────────────────────────── 셀 문구 만들기 ────────────────────────────

LINK = re.compile(r"\[([^\]]*)\]\((?!http)[^)]*\)")
BOLD = re.compile(r"\*\*([^*]+)\*\*")
CODE = re.compile(r"`([^`]*)`")


def plain(md: str) -> str:
    """마크다운 표기를 걷어낸다.

    문서 링크는 **글자만 남긴다** — 엑셀에서 `(../items.md)`는 따라갈 수 없는 흔적일 뿐이고,
    남겨 두면 X8이 금지한 "문서를 가리키는 칸"이 된다. 근거는 링크가 아니라 글로 넣는다.
    """
    s = LINK.sub(r"\1", md)
    s = BOLD.sub(r"\1", s)
    s = CODE.sub(r"\1", s)
    s = s.replace("〃", "위와 같음").replace("&nbsp;", " ")
    return re.sub(r"[ \t]+", " ", s).strip()


def joined(parts, sep=" · ") -> str:
    return sep.join(p for p in (x.strip() for x in parts) if p and p != "—")


# ──────────────────────────── 문서 가리키기를 걷어낸다 (X8) ────────────────────────────
#
# 마크다운에서 `assumptions.md §3.2`는 옳은 표기다 — 개발자는 그 파일을 열 수 있다.
# xlsx에서는 따라갈 수 없는 흔적이라 **무엇을 뜻하는지를 글로** 바꾼다. 규칙 §6.2가
# 정한 분담이다: 마크다운은 중복을 없애고, 펼치는 일은 생성기가 한다.
#
# **표에 없는 형태가 남으면 생성을 멈춘다.** 조용히 남기면 X8을 어긴 채 발표본이 된다.

POINTER_REWRITES = [
    # "상세는 …" 문장은 통째로 지운다 — 가리키는 곳이 xlsx에 없으니 앞 문장만 남으면 된다
    (r"\s*상세는[^.]*\.", ""),
    (r"\s*→\s*assumptions\.md\s*§[\d.]+", " (추정으로 등록해 둔 항목이다)"),
    (r"\(source-inconsistencies\.md\s*§[\d.]+\s*판독 결과\)", "(원문 그림 판독 결과)"),
    (r"source-inconsistencies\.md\s*§[\d.]+", "원문 모순·미정의 등록대장"),
    (r"두 축 정의는 README\s*§[\d.]+", "두 축(조작 축·범위 축)은 별도로 정의한다"),
    (r"README\s*§[\d.]+", "설계 명세 색인"),
    (r"저장소 내 유일한 수치 앵커\(unclear\.rule\.md\s*§\d+\s*예시\)", "저장소에 하나뿐인 수치 앵커"),
    (r"design-system\s*§[\d.]+", "디자인 시스템 문서"),
    (r"\(§6\.1의 \"[^\"]*\"\)", "(작성 규칙: 읽을 값이 없는 장식은 행으로 세지 않는다)"),
    (r"§4\.1 계측값", "계측 항목 표의 계측값"),
    (r"§5 이상 점수", "이상 점수 표의 값"),
    (r"§8 알람", "알람 표의 값"),
    (r"entities/equipment 4대 참조", "설비 4대 — 개수는 시연값"),
    # 남은 절 번호는 xlsx에서 가리킬 곳이 없다. 괄호째로 지우거나 번호만 지운다
    (r"\s*\(§[\d.]+\)", ""),
    (r"\s*—\s*§[\d.]+(?=\s*$)", ""),
    (r"§[\d.]+\s*", ""),
    (r"[^ ]*\.rule\.md\s*", "저장소 규칙 "),
    (r"[a-z-]+\.md\s*", ""),
]

POINTER_LEFT = re.compile(r"§|\.md\b|상세는|참조\b")


# 마크다운 강조가 셀에 남으면 발주처 양식에 `**✕**`이 그대로 찍힌다. 실제로 권한 열 59칸이
# 그랬다 — `plain()`을 거치지 않는 경로가 하나 있었다. 걷어냈는지를 생성기가 확인한다.
MARKUP_LEFT = re.compile(r"\*\*")


def for_xlsx(text: str) -> str:
    for pattern, replacement in POINTER_REWRITES:
        text = re.sub(pattern, replacement, text)
    return re.sub(r"\s{2,}", " ", text).strip(" ·")


# ──────────────────────────── 원천 파싱 ────────────────────────────

class Specs:
    def __init__(self) -> None:
        self.screens_md = read(SPECS / "screens.md")
        self.items_md = read(SPECS / "items.md")
        self.req_md = read(SPECS / "requirements.md")
        self.data_md = read(SPECS / "data-definition.md")
        self.screen_docs = {}
        for path in sorted(SCREENS.glob("SCR-*.md")):
            sid = path.name.split("-")[0] + "-" + path.name.split("-")[1] + "-" + path.name.split("-")[2]
            self.screen_docs[sid] = read(path)

        self.index = self._index()
        self.perms = self._perms()
        self.items = self._items()

    # screens.md §4·§4.1·§4.2 — 화면 목록
    def _index(self) -> list[dict]:
        out = []
        for header, rows in tables(self.screens_md):
            if header[:2] != ["#", "화면 ID"]:
                continue
            col = {name: k for k, name in enumerate(header)}
            for r in rows:
                sid = r[col["화면 ID"]]
                if not sid.startswith("SCR-"):
                    continue
                out.append(
                    {
                        "num": r[col["#"]],
                        "sid": sid,
                        "d1": plain(r[col["메뉴(1Depth)"]]),
                        "d2": plain(r[col["메뉴(2Depth)"]]),
                        "kind": plain(r[col["화면 구분"]]),
                        "desc": plain(r[col["설명"]]),
                        "req": plain(r[col["관련 요구사항"]]),
                    }
                )
        return out

    # screens.md §5 — 조작 축 권한
    def _perms(self) -> dict[str, tuple[str, str, str]]:
        found = table_with(self.screens_md, "화면 ID", "관리자", "운영자", "게스트")
        if not found:
            raise SystemExit("screens.md §5 권한 매트릭스를 못 찾았다")
        header, rows = found
        col = {name: k for k, name in enumerate(header)}
        out = {}
        for r in rows:
            # 매트릭스는 강조를 쓴다(`**✕**` — 닫힌 화면을 눈에 띄게). xlsx 셀에는 부호만 남긴다
            out[r[col["화면 ID"]]] = (
                plain(r[col["관리자"]]),
                plain(r[col["운영자"]]),
                plain(r[col["게스트"]]),
            )
        return out

    # items.md — 항목 사전
    def _items(self) -> dict[str, dict]:
        out: dict[str, dict] = {}
        prev_evidence = ""
        for header, rows in tables(self.items_md):
            if header[0] != "항목ID" or "근거" not in header:
                continue
            col = {name: k for k, name in enumerate(header)}
            for r in rows:
                ids = re.findall(r"`([A-Z]{2,5}-[A-Za-z0-9]+)`", r[0])
                if not ids:
                    continue
                evidence = plain(r[col["근거"]])
                if re.match(r"^(같음|〃|동일|위와 같음)", evidence):
                    evidence = prev_evidence + evidence.split(None, 1)[1] if False else prev_evidence
                else:
                    prev_evidence = evidence
                extra = joined(
                    [
                        f"{name} {plain(r[k])}"
                        for name, k in col.items()
                        if name not in ("항목ID", "라벨", "근거") and plain(r[k]) not in ("", "—")
                    ]
                )
                # 범위 표기(`SITE-01`~`SITE-10`)는 그 자체가 한 사전 행이다
                for item_id in ids:
                    out[item_id] = {
                        "id": item_id,
                        "set": item_id.split("-")[0],
                        "label": plain(r[col["라벨"]]) if "라벨" in col else plain(r[1]),
                        "evidence": evidence,
                        "extra": extra,
                        "range": len(ids) == 2 and "~" in r[0],
                        "row_ids": ids,
                    }
        return out

    # 화면 문서
    def components(self, sid: str) -> list[tuple[str, str]]:
        found = table_with(self.screen_docs[sid], "구성요소", "형태")
        if not found:
            return []
        header, rows = found
        col = {name: k for k, name in enumerate(header)}
        return [(plain(r[col["구성요소"]]), plain(r[col["형태"]])) for r in rows]

    def item_rows(self, sid: str) -> list[dict]:
        """`항목 목록` 절의 화면×항목 행. 소제목(#### …)이 영역 묶음을 나눈다."""
        body = section(self.screen_docs[sid], r"3(\.\d+)? 항목 목록")
        out = []
        group = ""
        lines = body.split("\n")
        for i, line in enumerate(lines):
            if line.startswith("#### "):
                group = plain(line[5:])
                continue
            s = line.strip()
            if not s.startswith("|"):
                continue
            cells = split_row(s)
            if cells[:1] == ["항목ID"] or SEPARATOR.match(s):
                continue
            if len(cells) < 5:
                continue
            out.append(
                {
                    "ids": re.findall(r"`([A-Z]{2,5}-[A-Za-z0-9]+)`", cells[0]),
                    "raw_id": plain(cells[0]),
                    "area": plain(cells[1]),
                    "name": plain(cells[2]),
                    "form": plain(cells[3]),
                    # 사전에 없는 항목의 근거는 이 칸에만 있다. 사전 항목은 `—`이고 사전에서 가져온다
                    "evidence": plain(cells[4]),
                    "group": group,
                }
            )
        return out

    def actions(self, sid: str) -> dict[str, str]:
        found = table_with(self.screen_docs[sid], "xlsx 열", "값")
        out = {"SUBMIT": "", "CANCEL": "", "After Action": ""}
        if found:
            for r in found[1]:
                if r[0] in out:
                    out[r[0]] = plain(r[1])
        return out

    def exceptions(self, sid: str) -> str:
        body = section(self.screen_docs[sid], r"6\. 상태와 예외")
        found = table_with(body, "상황", "표시")
        if not found:
            return ""
        return " / ".join(f"{plain(r[0])} → {plain(r[1])}" for r in found[1])

    def rationale(self, sid: str) -> list[str]:
        return self._paragraphs(section(self.screen_docs[sid], r"7\. 설계 근거"))

    def open_items(self, sid: str) -> list[str]:
        body = section(self.screen_docs[sid], r"8\. 미정·제약")
        out = self._paragraphs(body)
        found = table_with(body, "항목") or table_with(body, "무엇")
        if found:
            out += [joined([plain(c) for c in r]) for r in found[1]]
        return out

    @staticmethod
    def _paragraphs(body: str) -> list[str]:
        out = []
        for chunk in re.split(r"\n\s*\n", body):
            t = chunk.strip()
            if not t or t.startswith("|") or t.startswith("#") or t.startswith("---"):
                continue
            out.append(plain(t.replace("\n", " ")))
        return out

    def data_rows(self, sid: str) -> list[dict]:
        """화면 문서 §4 요구 데이터."""
        body = section(self.screen_docs[sid], r"4\. 요구 데이터")
        out = []
        for header, rows in tables(body):
            if header[0] != "데이터":
                continue
            col = {name: k for k, name in enumerate(header)}
            for r in rows:
                out.append({name: plain(r[k]) for name, k in col.items()})
        return out


# ──────────────────────────── xlsx 쓰기 ────────────────────────────

def find_header_row(ws, *must: str) -> int:
    """헤더 행을 **문자열로 찾는다.** 요구사항정의서는 시트마다 3행/4행으로 다르다."""
    for row in ws.iter_rows(min_row=1, max_row=12):
        values = [str(c.value).replace("\n", " ") if c.value is not None else "" for c in row]
        if all(any(want in v for v in values) for want in must):
            return row[0].row
    raise SystemExit(f"[{ws.title}] 헤더 행을 못 찾았다: {must}")


def header_map(ws, header_row: int) -> dict[str, int]:
    out = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        key = re.sub(r"\s+", " ", str(cell.value).replace("\n", " ")).strip()
        out[key] = cell.column
    return out


def capture_style(ws, row: int, max_col: int) -> dict[int, object]:
    return {c: copy(ws.cell(row=row, column=c)._style) for c in range(1, max_col + 1)}


def clear_below(ws, header_row: int, max_col: int) -> None:
    """헤더 아래의 값을 지운다. 서식·병합은 건드리지 않는다."""
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.value = None


LEFTOVERS: list[str] = []
MARKUP: list[str] = []


def flag(sheet: str, coord: str, value: str) -> None:
    """셀에 남으면 안 되는 것을 모은다. 실행 끝에 하나라도 있으면 실패로 끝낸다."""
    if POINTER_LEFT.search(value):
        LEFTOVERS.append(f"[{sheet}] {coord} — {value[:90]}")
    if MARKUP_LEFT.search(value):
        MARKUP.append(f"[{sheet}] {coord} — {value[:90]}")


def write_rows(ws, header_row: int, style: dict, rows: list[list], max_col: int) -> None:
    for offset, values in enumerate(rows):
        r = header_row + 1 + offset
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if c in style:
                cell._style = copy(style[c])
            value = values[c - 1] if c - 1 < len(values) else None
            if isinstance(value, str):
                value = for_xlsx(value)
                flag(ws.title, cell.coordinate, value)
            cell.value = value
        # 줄바꿈 셀은 높이를 엑셀이 정하게 둔다 — 템플릿 고정 높이를 물려받으면 글이 잘린다
        ws.row_dimensions[r].height = None


def drop_column(ws, col: int) -> None:
    ws.delete_cols(col)


# ──────────────────────────── 화면설계서 ────────────────────────────

SHEET_FOR = {"AD": "관리자용", "OP": "운영자용", "GU": "게스트용"}
NEW_COLUMNS = ["근거", "비고"]


def screen_evidence(sp: Specs, item: dict) -> str:
    """항목 사전의 근거를 **한 칸에 글로** 펼친다(X8).

    마크다운은 사전 항목의 근거를 되풀이하지 않고 항목ID로 가리킨다(E1). 심사자는 그
    항목ID만 적힌 칸을 보고 뜻을 알 수 없으므로, 펼치는 일을 생성기가 한다.
    사전에 없는 항목은 화면 문서의 `근거` 칸이 유일한 출처다.
    """
    parts = []
    for item_id in item["ids"]:
        entry = sp.items.get(item_id)
        if entry:
            parts.append(f"{entry['label']}: {entry['evidence']}")
    own = item.get("evidence", "")
    if own and own != "—":
        parts.append(own)
    # 근거가 `형태` 칸 안에 적힌 행이 있다(`20~30%` `[원문 p.27·31]`). 그 태그를 근거 열로 끌어온다 —
    # 옆 칸에 있으니 xlsx는 자립하지만, 근거 열이 비면 근거가 없는 행처럼 보인다
    if not parts:
        tags = tags_in(item["form"])
        if tags:
            parts.append(f"{tags} — 형태 열에 값과 함께 적었다")
    seen, uniq = set(), []
    for p in parts:
        if p not in seen:
            seen.add(p)
            uniq.append(p)
    return " / ".join(uniq)


def build_screen_design(sp: Specs) -> None:
    path = XLSX["screen"]
    wb = openpyxl.load_workbook(path)

    by_sheet: dict[str, list[list]] = {name: [] for name in SHEET_FOR.values()}
    appendix_items: list[list] = []
    appendix_reason: list[list] = []

    for entry in sp.index:
        sid = entry["sid"]
        role = sid.split("-")[1]
        targets = list(SHEET_FOR.values()) if role == "CO" else [SHEET_FOR[role]]
        perm = sp.perms.get(sid, ("", "", ""))
        actions = sp.actions(sid)
        components = sp.components(sid)
        exceptions = sp.exceptions(sid)
        items = sp.item_rows(sid)
        opens = sp.open_items(sid)

        summary = [
            entry["num"],
            sid,
            entry["d1"],
            entry["d2"] or "—",
            entry["kind"],
            entry["desc"],
            perm[0],
            perm[1],
            perm[2],
            joined([c[0] for c in components], ", "),
            joined([c[1] for c in components], ", "),
            actions["SUBMIT"] or "N",
            actions["CANCEL"] or "N",
            actions["After Action"] or "—",
            f"관련 요구사항 {entry['req']}" if entry["req"] else "",
            exceptions,
        ]
        base = entry["num"].split(".")[0]
        rows = [summary]
        for k, item in enumerate(items, start=1):
            rows.append(
                [
                    f"{base}.{k}",
                    sid,
                    entry["d1"],
                    "—",
                    entry["kind"],
                    joined([item["group"], item["area"]], " > "),
                    perm[0],
                    perm[1],
                    perm[2],
                    item["name"],
                    item["form"],
                    "",
                    "",
                    "",
                    screen_evidence(sp, item),
                    "",
                ]
            )
        for sheet in targets:
            by_sheet[sheet] += rows

        for text in sp.rationale(sid):
            appendix_reason.append([sid, "근거", text, tags_in(text)])
        for text in opens:
            appendix_reason.append([sid, "미정", text, tags_in(text)])

    for item_id, entry in sorted(sp.items.items()):
        used = [
            e["sid"]
            for e in sp.index
            if any(item_id in row["ids"] for row in sp.item_rows(e["sid"]))
        ]
        appendix_items.append(
            [
                item_id,
                entry["set"],
                entry["label"],
                entry["extra"],
                entry["evidence"],
                ", ".join(sorted(set(used))) or "—",
            ]
        )

    for sheet, rows in by_sheet.items():
        ws = wb[sheet]
        header_row = find_header_row(ws, "화면 ID", "구성요소", "After Action")
        cols = header_map(ws, header_row)
        if "PC" in cols:
            drop_column(ws, cols["PC"])
            cols = header_map(ws, header_row)
        style = capture_style(ws, header_row + 1, 14)
        header_style = copy(ws.cell(row=header_row, column=14)._style)
        for offset, name in enumerate(NEW_COLUMNS):
            col = 15 + offset
            if name not in cols:
                cell = ws.cell(row=header_row, column=col)
                cell.value = name
                cell._style = copy(header_style)
                ws.column_dimensions[get_column_letter(col)].width = 60
                style[col] = copy(style[14])
        clear_below(ws, header_row, 16)
        write_rows(ws, header_row, style, rows, 16)
        print(f"  [{sheet}] {len(rows)}행")

    add_appendix(
        wb,
        "부록. 항목 목록",
        ["항목ID", "집합", "라벨", "단위·자릿수·범위", "근거", "쓰이는 화면"],
        appendix_items,
        wb["부록. 화면 구성요소 목록"],
    )
    add_appendix(
        wb,
        "부록. 설계 근거",
        ["화면 ID", "구분", "내용", "근거 태그"],
        appendix_reason,
        wb["부록. 화면 구성요소 목록"],
    )
    stamp_history(wb, "화면설계서")
    wb.save(path)
    print(f"  → {path.name}")


"""대괄호가 열린 뒤의 **머리말만** 본다.

`[PROVISIONAL, TBD-31]`처럼 한 괄호에 둘을 적은 곳이 있어 `\\]`로 닫으면 놓친다 —
실제로 두 행이 그렇게 근거 없는 행으로 보였다.
"""
TAG = re.compile(r"\[(원문 발표|원문|공정자료|데이터셋|파생|PROVISIONAL|TBD|INC-\d{2,3}|설계|회의|사용자)")


def tags_in(text: str) -> str:
    """그 문장이 업은 근거 태그만 뽑아 따로 보여준다 — 심사자가 근거의 종류를 훑을 수 있게."""
    found = []
    for m in TAG.finditer(text):
        if m.group(1) not in found:
            found.append(m.group(1))
    # `E3를 지킨다`처럼 조사가 붙으면 `\bE3\b`가 안 걸린다 — 뒤는 숫자만 아니면 된다
    for m in re.finditer(r"\bE([1-6])(?![0-9])", text):
        if f"E{m.group(1)}" not in found:
            found.append(f"E{m.group(1)}")
    return " · ".join(found)


def add_appendix(wb, title: str, header: list[str], rows: list[list], model) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    head_style = copy(model.cell(row=1, column=1)._style)
    body_style = copy(model.cell(row=2, column=1)._style)
    widths = [16, 8, 26, 34, 76, 30]
    for c, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c)
        cell.value = name
        cell._style = copy(head_style)
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1] if c - 1 < len(widths) else 40
    for r, values in enumerate(rows, start=2):
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=r, column=c)
            cell._style = copy(body_style)
            value = values[c - 1] if c - 1 < len(values) else None
            if isinstance(value, str):
                value = for_xlsx(value)
                flag(title, cell.coordinate, value)
            cell.value = value
    ws.freeze_panes = "A2"
    print(f"  [{title}] {len(rows)}행")


# ──────────────────────────── 요구사항정의서 ────────────────────────────

REQ_SHEETS = {
    "CO": "시스템 공통",
    "AD": "관리자",
    "OP": "운영자",
    "GU": "게스트",
    "NF": "비기능요구사항",
}


def build_requirements(sp: Specs) -> None:
    path = XLSX["req"]
    wb = openpyxl.load_workbook(path)

    buckets: dict[str, list[list]] = {name: [] for name in REQ_SHEETS.values()}
    for header, rows in tables(sp.req_md):
        if header[0] != "요구사항ID" or "상세설명" not in header:
            continue
        col = {name: k for k, name in enumerate(header)}
        for r in rows:
            rid = r[col["요구사항ID"]]
            m = re.match(r"REQ-(CO|AD|OP|GU|NF)-\d{3}", rid)
            if not m:
                continue
            sheet = REQ_SHEETS[m.group(1)]
            detail = plain(r[col["상세설명"]])
            applied = plain(r[col["적용방안 및 제약사항"]])
            accepted = r[col["수용"]].strip() if "수용" in col else ""
            buckets[sheet].append(
                [
                    plain(r[col["대분류"]]),
                    plain(r[col["중분류"]]),
                    rid,
                    plain(r[col["요구사항명"]]),
                    detail,
                    applied,
                    "비기능" if m.group(1) == "NF" else "기능",
                    plain(r[col["우선순위"]]),
                    "추가",
                    "신규 작성",
                    accepted or accept_from(applied),
                    plain(r[col["관련"]]),
                    plain(r[col["화면·상태"]]),
                ]
            )

    for sheet, rows in buckets.items():
        ws = wb[sheet]
        header_row = find_header_row(ws, "요구사항ID", "변경구분", "수용여부")
        style = capture_style(ws, header_row + 1, 13)
        clear_below(ws, header_row, 13)
        write_rows(ws, header_row, style, rows, 13)
        print(f"  [{sheet}] {len(rows)}행")

    for name in ("1", "2", "3"):
        if name in wb.sheetnames:
            del wb[name]
            print(f"  [{name}] 빈 시트 삭제")
    stamp_history(wb, "요구사항정의서")
    wb.save(path)
    print(f"  → {path.name}")


def accept_from(applied: str) -> str:
    """`수용` 열이 없을 때 구현 상태로 정한다(규칙 §5.5). 대상아님만 N이다."""
    return "N" if "대상아님" in applied else "Y"


# ──────────────────────────── 데이터 정의서 ────────────────────────────

def build_data_definition(sp: Specs) -> None:
    path = XLSX["data"]
    wb = openpyxl.load_workbook(path)
    ws = wb["데이터 정의서"]

    ui_place = {}
    for entry in sp.index:
        for row in sp.data_rows(entry["sid"]):
            ui_place[(entry["sid"], row.get("데이터", ""))] = row.get("UI 위치", "")

    rows: list[list] = []
    group = ""
    for line in sp.data_md.split("\n"):
        m = re.match(r"^#{2,3} (?:\d+(?:\.\d+)?\.? )?(.+)$", line)
        if m and not line.startswith("### 변경"):
            group = plain(m.group(1))
            # 절 제목의 괄호는 타입 이름이다(`예측 (ForecastPoint · ForecastSummary)`).
            # 그룹 이름은 도메인 이름이어야 한다 — 타입은 `데이터 명` 열이 이미 말한다
            group = re.sub(r"\s*\([^)]*\)\s*$", "", group).strip()
        if not line.strip().startswith("|"):
            continue
        cells = split_row(line)
        if len(cells) != 9 or cells[0] in ("요구사항ID", "") or SEPARATOR.match(line.strip()):
            continue
        rid, sid, name, typ, unit, calc, source, ui, place = [plain(c) for c in cells]
        if not (rid.startswith("REQ-") or rid == "—"):
            continue
        rows.append(
            [
                rid,
                sid,
                group,
                name,
                typ,
                unit,
                calc,
                source,
                ui,
                place or ui_place.get((sid, name), ""),
                note_from(source, calc),
            ]
        )

    header_row = find_header_row(ws, "데이터 명", "수집 출처/레퍼런스")
    style = capture_style(ws, header_row + 1, 11)
    clear_below(ws, header_row, 11)
    write_rows(ws, header_row, style, rows, 11)
    print(f"  [데이터 정의서] {len(rows)}행")
    stamp_history(wb, "데이터 정의서")
    wb.save(path)
    print(f"  → {path.name}")


def note_from(source: str, calc: str) -> str:
    """비고는 **미확정 사유**를 글로 적는 칸이다(X2·X8)."""
    notes = []
    if "PROVISIONAL" in source or "PROVISIONAL" in calc:
        notes.append("임시값 — 확정 시 src/shared/config/provisional.ts 한 파일을 교체한다")
    for m in re.finditer(r"TBD-\d{2}|TBD", source + " " + calc):
        notes.append(f"{m.group(0)} 원문 미확정")
        break
    for m in re.finditer(r"INC-\d{2,3}", source + " " + calc):
        notes.append(f"{m.group(0)} 원문 모순 — 채택안은 출처 열에 적었다")
        break
    return " / ".join(dict.fromkeys(notes))


# ──────────────────────────── 변경이력 ────────────────────────────

def stamp_history(wb, doc_name: str) -> None:
    """`변경이력` 시트의 작성일자를 갱신한다. 버전은 v0.1을 유지한다(규칙 §7.3)."""
    ws = wb["변경이력"]
    for row in ws.iter_rows(min_row=1, max_row=12):
        for cell in row:
            if cell.value == "작성일자":
                ws.cell(row=cell.row, column=cell.column + 1).value = BUILD_DATE
    for row in ws.iter_rows(min_row=8, max_row=20):
        if row[0].value not in (None, ""):
            row[0].value = BUILD_DATE
            break


# ──────────────────────────── 실행 ────────────────────────────

def main() -> int:
    for path in XLSX.values():
        if not path.exists():
            raise SystemExit(f"원본이 없다: {path}")
        shutil.copy2(path, path.with_suffix(".xlsx.bak"))
    print(f"백업 3건 · 생성일 {BUILD_DATE}")

    sp = Specs()
    print(f"화면 {len(sp.index)}개 · 항목 사전 {len(sp.items)}개")

    print("화면설계서")
    build_screen_design(sp)
    print("요구사항정의서")
    build_requirements(sp)
    print("데이터 정의서")
    build_data_definition(sp)

    failed = False
    if LEFTOVERS:
        print("")
        print("X8 위반 %d건 — 문서를 가리키는 칸이 남았다" % len(LEFTOVERS))
        for line in LEFTOVERS[:40]:
            print("  " + line)
        print("POINTER_REWRITES에 그 형태를 등록하고 다시 실행한다.")
        failed = True
    if MARKUP:
        print("")
        print("마크다운 강조가 남은 칸 %d건 — plain()을 거치지 않는 경로가 있다" % len(MARKUP))
        for line in MARKUP[:40]:
            print("  " + line)
        failed = True
    if failed:
        return 1
    print("")
    print("X8 통과 · 마크다운 잔재 없음")
    return 0


if __name__ == "__main__":
    sys.exit(main())
